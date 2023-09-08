from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pathlib
import json

mitre = []
mitrenav = {"name": "layer",	"versions": {		"attack": "13",		"navigator": "4.8.2",		"layer": "4.4"	},	"domain": "enterprise-attack",	"description": "",	"sorting": 0,	"layout": {		"layout": "side",		"aggregateFunction": "average",		"showID": "false",		"showName": "true",		"showAggregateScores": "false",		"countUnscored": "false"	},	"hideDisabled": "false",	"techniques": ""    }

file = input('Enter filename\n Example: C:/Users/richard/excel.xlsx\n')
xls = load_workbook(filename=file)

sheets_excel  = input('Your excel has more than one sheet? Yes or No\n')

if sheets_excel == 'yes' or sheets_excel == 'Yes':
    name_sheet = input('What is the name Excel sheet?\n')
    print(name_sheet)
    sheet = xls[name_sheet] #set sheet
    last_column = len(list(sheet.columns))
    last_row = len(list(sheet.rows))
    print("Filas: "+str(last_row)+" Columnas: "+str(last_column))
else:
    sheet = xls.active 
    last_column = len(list(sheet.columns))
    last_row = len(list(sheet.rows))

columtactica = input('What is the excel column that store tactics?\n')
columtecnica = input('What is the column that store techniques?\n')

tecnicas = []
for row in range(2, last_row + 1): #identify techinique from xlsx and add to list
    if sheet[columtactica + str(row)].value != None and sheet[columtecnica + str(row)].value != None:
        tecnicas.append(sheet[columtecnica + str(row)].value)

#set list with technique, tactic and color based on UC's technique amount.
for row in range(2, last_row + 1):
    my_dict = {}
    if sheet[columtactica + str(row)].value != None and sheet[columtecnica + str(row)].value != None:
        if tecnicas.count(sheet[columtecnica + str(row)].value) < 10:
            my_dict["color"] = "#A3E4D7"
        elif tecnicas.count(sheet[columtecnica + str(row)].value) > 10 and tecnicas.count(sheet[columtecnica + str(row)].value) < 30:
            my_dict["color"] = "#76D7C4"
        elif tecnicas.count(sheet[columtecnica + str(row)].value) > 30 and tecnicas.count(sheet[columtecnica + str(row)].value) < 50:
            my_dict["color"] = "#48C9B0"
        elif tecnicas.count(sheet[columtecnica + str(row)].value) > 50:
            my_dict["color"] = "#1ABC9C"
        my_dict[sheet[columtactica + str(1)].value] = sheet[columtactica + str(row)].value 
        #key: value --> key=columtactica's first row, value=columtactica's for row
        my_dict[sheet[columtecnica + str(1)].value] = sheet[columtecnica + str(row)].value
        #key: value --> key=columtecnica's first row, value=columtecnica's for row
        mitre.append(my_dict) #append each dictionary completed to a list    

mitrenav["techniques"] = mitre #adding whole tacticas in key "technique"

pathcurrentdir = pathlib.Path(__file__).parent.resolve()
data = json.dumps(mitrenav, sort_keys=True, indent=4)
with open(str(pathcurrentdir)+'/MitreNavigator.json', 'w', encoding='utf-8') as f:
    f.write(data)
